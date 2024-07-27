import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import StaleElementReferenceException, TimeoutException, NoSuchElementException
from PIL import Image
import pytesseract
import openpyxl
import tkinter.messagebox as messagebox
import threading

# Defina o caminho para o executável do Tesseract
pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'

def esperar_elemento(driver, by, value, timeout=10):
    """Função para aguardar a presença de um elemento na página."""
    wait = WebDriverWait(driver, timeout)
    return wait.until(EC.presence_of_element_located((by, value)))

def ler_dados_excel(arquivo_excel):
    try:
        workbook = openpyxl.load_workbook(arquivo_excel)
        sheet = workbook.active

        # Aqui você pode fazer o que quiser com os dados da planilha, por exemplo, imprimir cada célula
        for row in sheet.iter_rows(values_only=True):
            print(row)
            

        workbook.close()

    except Exception as e:
        print(f"Erro ao ler dados do arquivo Excel: {str(e)}")

def capturar_e_preencher_captcha(tipo, login, senha, arquivo_excel):
    url_mapping = {
        'Cajamar': 'https://cajamar.rf1consig.com.br/SGConsignataria/ConsigAcessoUsuarioLogar.aspx',
        'Boavista': 'https://boavista.rf1consig.com.br/SGConsignataria/ConsigAcessoUsuarioLogar.aspx',
        'Juruti': 'https://juruti.rf1consig.com.br/SGConsignataria/ConsigAcessoUsuarioLogar.aspx',
        'Portel': 'https://portel.rf1consig.com.br/SGConsignataria/ConsigAcessoUsuarioLogar.aspx',
        'Caxias': 'https://campos.rf1consig.com.br/SGConsignataria/ConsigAcessoUsuarioLogar.aspx'

        # Adicione outros tipos e URLs aqui conforme necessário
    }

    if tipo not in url_mapping:
        print("Tipo inválido.")
        return

    try:
        options = webdriver.FirefoxOptions()
        options.headless = False
        driver = webdriver.Firefox(options=options)
        driver.get(url_mapping[tipo])

        time.sleep(2)

        campo_login = driver.find_element(By.ID, "ctl00_ContentPlaceHolder1_txtUsuario")
        campo_login.send_keys(login)

        time.sleep(2)

        campo_senha = driver.find_element(By.ID, "ctl00_ContentPlaceHolder1_txtSenha")
        campo_senha.send_keys(senha)

        time.sleep(2)

        campo_captcha = driver.find_element(By.ID, "ctl00_ContentPlaceHolder1_txtValidaCaptcha")

        if not campo_captcha.get_attribute("value"):
            # Esperar até que a imagem do captcha seja carregada
            try:
                captcha_element = esperar_elemento(driver, By.CSS_SELECTOR, "img[src='Captcha.aspx']")
            except TimeoutException:
                print("Elemento do captcha não encontrado.")
                driver.quit()
                return

            x = captcha_element.location['x']
            y = captcha_element.location['y']
            width = captcha_element.size['width']
            height = captcha_element.size['height']

            # Tirar uma captura de tela da página inteira
            screenshot_path = "pagina_screenshot.png"
            driver.save_screenshot(screenshot_path)

            time.sleep(3)

            # Definir as coordenadas e dimensões da área que você deseja recortar
            # Ler as coordenadas do arquivo
            with open('coordenadas.txt', 'r') as f:
                x, y, width, height = map(int, f.readline().split(','))

            # Abrir a imagem usando Pillow
            screenshot = Image.open("pagina_screenshot.png")

            # Recortar a área desejada
            captcha_area = screenshot.crop((x, y, x + width, y + height))

            time.sleep(3)

            captcha_area.save("captcha_area.png")

            # Utilizar OCR para ler o texto do captcha
            texto_captcha = pytesseract.image_to_string(captcha_area)
            print("Texto do Captcha:", texto_captcha)

            campo_captcha.send_keys(texto_captcha)

            time.sleep(2)

            if not campo_captcha.get_attribute("value"):
                print("Campo de captcha ainda está vazio. Fechando a página e recarregando.")
                driver.close()
                return capturar_e_preencher_captcha(tipo, login, senha, arquivo_excel)

        btn_entrar = driver.find_element(By.ID, "ctl00_ContentPlaceHolder1_btnEntrar")
        btn_entrar.click()

        time.sleep(10)

        mensagem_codigo_invalido = driver.find_elements(By.XPATH, "//td[@id='SGConsigMensagem_TD_Tabela' and contains(text(), 'Código de segurança inválido')]")
        if mensagem_codigo_invalido:
            print("Código de segurança inválido.")
            driver.quit()
            return capturar_e_preencher_captcha(tipo, login, senha, arquivo_excel)

        mensagem_usuario_logado = driver.find_elements(By.XPATH, "//td[@id='SGConsigMensagem_TD_Tabela' and contains(text(), 'O usuário') and contains(text(), 'já se encontra logado')]")
        if mensagem_usuario_logado:
            print("Usuário já está logado.")
            driver.quit()
            return

        time.sleep(2)

        # Clicar no botão "GESTOR"
        btn_gestor = driver.find_element(By.XPATH, "//a[@href='GESTOR/CADPessoaListar.aspx' and @class='btn btn-menu']")
        btn_gestor.click()

        try:
            btn_consultar = driver.find_element(By.ID, "ctl00_ctl00_ContentPlaceHolder1_ContentPlaceHolder1_btnListar")
        except NoSuchElementException:
            time.sleep(2)
            driver.back()
            time.sleep(1)
            btn_gestor = driver.find_element(By.XPATH, "//a[@href='GESTOR/CADPessoaListar.aspx' and @class='btn btn-menu']")
            btn_gestor.click()

        time.sleep(2)

        # Carregar o arquivo Excel
        wb = openpyxl.load_workbook(arquivo_excel)
        sheet = wb.active

        # Contar quantas células estão preenchidas na primeira coluna a partir da segunda linha
        total_linhas_primeira_coluna = sum(1 for cell in sheet['A'][1:] if cell.value is not None)

        # Contar quantas células estão preenchidas na terceira coluna a partir da segunda linha
        total_linhas_terceira_coluna = sum(1 for cell in sheet['C'][1:] if cell.value is not None)

        # Determinar a linha inicial para o loop
        linha_inicial_loop = (total_linhas_terceira_coluna) + 2

        # Determinar total linha coluna 1

        total_linhas_coluna_completa = (total_linhas_primeira_coluna) + 1

        # Definindo o tempo de início do loop
        tempo_inicio = time.time()

        # Loop para processar cada linha do arquivo Excel a partir da linha inicial determinada
        row = linha_inicial_loop
        while row <= total_linhas_coluna_completa:

            matricula = sheet.cell(row=row, column=1).value

            # Verificar se passaram 3 horas
            tempo_atual = time.time()
            if tempo_atual - tempo_inicio >= 10800:  # 10800 segundos são equivalentes a 3 horas
                # Encontrar o elemento "Sair" e clicar nele
                link_sair = driver.find_element(By.ID, "ctl00_ctl00_ContentPlaceHolder1_lnkSair")
                link_sair.click()
                time.sleep(2)  # Esperar 2 segundos

                # Encerrar o driver
                driver.quit()

                # Retornar da função
                return capturar_e_preencher_captcha(tipo, login, senha, arquivo_excel)

            # Seu código existente aqui (continuação do loop)

            # Preencher campos de matrícula
            campo_matricula = None
            while campo_matricula is None:
                try:
                    time.sleep(2)
                    campo_matricula = driver.find_element(By.ID, "ctl00_ctl00_ContentPlaceHolder1_ContentPlaceHolder1_txtCPF")
                except StaleElementReferenceException:
                    time.sleep(2)
                    print("Elemento de matrícula obsoleto. Tentando novamente...")
                    time.sleep(1)

            time.sleep(2)
            driver.execute_script("arguments[0].value = '';", campo_matricula)
            campo_matricula.send_keys(matricula)

            # Clicar no botão de consulta
            btn_consultar = driver.find_element(By.ID, "ctl00_ctl00_ContentPlaceHolder1_ContentPlaceHolder1_btnListar")
            time.sleep(3)
            btn_consultar.click()

            # Verificar se nenhum registro foi encontrado
            try:
                mensagem = driver.find_element(By.ID, "SGConsigMensagem_TD_Tabela").text
                time.sleep(1)
                if "+ Nenhum registro encontrado." in mensagem:
                    # Inserir valor 0.00 na coluna 3 da mesma linha
                    sheet.cell(row=row, column=3).value = "Não Tem"
                    wb.save(arquivo_excel)
                    row += 1  # Avança para a próxima linha
                    continue
            except:
                pass

            time.sleep(3)

            # Localize o elemento span pelo ID
            element = driver.find_element(By.ID, "ctl00_ctl00_ContentPlaceHolder1_ContentPlaceHolder1_lblCartaoAdianmento")

            # Obtenha o texto do elemento usando o atributo innerHTML
            valor_consulta = element.get_attribute("innerHTML")

            # Agora você tem o texto "R$ 711,34" armazenado na variável valor_consulta
            print("Valor da consulta:", valor_consulta)

            time.sleep(1)

            # Preencher coluna 3 com o valor da consulta
            sheet.cell(row=row, column=3).value = valor_consulta

            # Localizar Margem Consignavel Atual

            time.sleep(1)

            element = driver.find_element(By.ID, "ctl00_ctl00_ContentPlaceHolder1_ContentPlaceHolder1_lblMargemConsignavel")

            # Obtenha o texto do elemento usando o atributo innerHTML
            valor_consultaatual = element.get_attribute("innerHTML")

            # Agora você tem o texto "R$ 711,34" armazenado na variável valor_consulta
            print("Valor da consulta Atual:", valor_consultaatual)

            time.sleep(1)

            # Preencher coluna 3 com o valor da consulta
            sheet.cell(row=row, column=4).value = valor_consultaatual

            wb.save(arquivo_excel)

            # Desativar a atualização automática da página
            driver.execute_script("window.onbeforeunload = function() {};")

            row += 1  # Avança para a próxima linha

        # Mensagem de conclusão
        print("Processo finalizado.")
        messagebox.showinfo("Concluído", "Processo do arquivo Excel concluído com sucesso!")

    except Exception as e:
        print(f"Ocorreu um erro: {str(e)}")

# Exemplo de uso:
# capturar_e_preencher_captcha('Cajamar', 'seu_login', 'sua_senha', 'caminho_para_arquivo_excel')
